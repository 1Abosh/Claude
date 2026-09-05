import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
BLUE = Font(name=FONT_NAME, size=10, color="0000FF")
BLACK = Font(name=FONT_NAME, size=10, color="000000")
GREEN = Font(name=FONT_NAME, size=10, color="008000")
BOLD = Font(name=FONT_NAME, size=10, bold=True)
TITLE = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFF")
SECTION = Font(name=FONT_NAME, size=11, bold=True, color="1F4E78")
NOTE = Font(name=FONT_NAME, size=9, italic=True, color="666666")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
INPUT_FILL = PatternFill("solid", fgColor="FFFDE7")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()
wb.remove(wb.active)

registry = {}  # symbol -> "'Sheet Name'!$B$12"

def qname(sheet_name):
    return f"'{sheet_name}'" if " " in sheet_name or "(" in sheet_name else sheet_name

class Builder:
    def __init__(self, ws, name):
        self.ws = ws
        self.name = name
        self.row = 1

    def title(self, text):
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=5)
        c = self.ws.cell(self.row, 1, text)
        c.font = TITLE
        c.fill = TITLE_FILL
        c.alignment = Alignment(vertical="center")
        self.ws.row_dimensions[self.row].height = 22
        self.row += 2

    def section(self, text):
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=5)
        c = self.ws.cell(self.row, 1, text)
        c.font = SECTION
        c.fill = SECTION_FILL
        self.row += 1

    def colheads(self):
        heads = ["Parameter", "Value", "Units", "Notes / Source"]
        for i, h in enumerate(heads, start=1):
            c = self.ws.cell(self.row, i, h)
            c.font = BOLD
        self.row += 1

    def kv(self, label, value, key=None, units="", note="", is_input=False, is_link=False,
           number_format=None, bold_value=False):
        r = self.row
        self.ws.cell(r, 1, label).font = BLACK
        vcell = self.ws.cell(r, 2, value)
        if is_input:
            vcell.font = BLUE
            vcell.fill = INPUT_FILL
        elif is_link:
            vcell.font = GREEN
        else:
            vcell.font = Font(name=FONT_NAME, size=10, bold=bold_value)
        if number_format:
            vcell.number_format = number_format
        self.ws.cell(r, 3, units).font = NOTE
        self.ws.cell(r, 4, note).font = NOTE
        for col in (1, 2, 3, 4):
            self.ws.cell(r, col).border = BORDER
        if key:
            registry[key] = f"{qname(self.name)}!$B${r}"
        self.row += 1
        return f"{qname(self.name)}!$B${r}"

    def blank(self, n=1):
        self.row += n

    def widths(self, widths=(34, 16, 10, 55)):
        for i, w in enumerate(widths, start=1):
            self.ws.column_dimensions[get_column_letter(i)].width = w


# ======================================================================
# SHEET 1: INPUTS
# ======================================================================
ws = wb.create_sheet("Inputs")
b = Builder(ws, "Inputs")
b.widths()
b.title("SHELL & TUBE HEAT EXCHANGER — KERN'S METHOD (Coulson & Richardson Vol.6, Ch.12)")
ws.cell(3, 1, "Excel port of CIO_Part_1.ipynb. Blue/yellow cells are editable inputs; everything else is a live formula.").font = NOTE
b.row = 5

b.section("HOT FLUID — SHELL SIDE  (MIL-PRF-23699 turbine oil)")
b.colheads()
b.kv("Hot fluid inlet temperature, Th_in", 80, "Th_in", "°C", "Notebook default", is_input=True)
b.kv("Hot fluid outlet temperature, Th_out", 45, "Th_out", "°C", "Notebook default", is_input=True)
b.kv("Hot fluid mass flow rate, m_dot_h", "=10000/3600", "m_dot_h", "kg/s", "10,000 kg/h", is_input=True)
b.kv("Hot fluid density, rho_h", 960.08, "rho_h", "kg/m³", "MIL-PRF-23699 @ bulk T (const. in model)", is_input=True)
b.kv("Hot fluid specific heat, cp_h", 1903.1, "cp_h", "J/kg·K", "MIL-PRF-23699", is_input=True)
b.kv("Hot fluid thermal conductivity, k_h", 0.1402, "k_h", "W/m·K", "MIL-PRF-23699", is_input=True)
b.kv("Hot fluid viscosity, mu_h", 0.0116192, "mu_h", "Pa·s", "MIL-PRF-23699", is_input=True)
b.blank()

b.section("COLD FLUID — TUBE SIDE  (Water)")
b.colheads()
b.kv("Cold fluid inlet temperature, Tc_in", 20, "Tc_in", "°C", "Notebook default (Cell 1)", is_input=True)
b.kv("Cold fluid outlet temperature, Tc_out", 38, "Tc_out", "°C", "Notebook default (Cell 1)", is_input=True)
b.kv("Cold fluid specific heat, cp_c", 4183, "cp_c", "J/kg·K", "Water", is_input=True)
b.kv("Cold fluid density, rho_c", 985.69, "rho_c", "kg/m³", "Water", is_input=True)
b.kv("Cold fluid viscosity, mu_c", 0.00050362, "mu_c", "Pa·s", "Water (503.62e-6)", is_input=True)
b.kv("Cold fluid thermal conductivity, k_c", 0.64601, "k_c", "W/m·K", "Water", is_input=True)
b.blank()

b.section("MATERIALS / FOULING")
b.colheads()
b.kv("Tube wall conductivity, k_wall", 45.0, "k_wall", "W/m·K", "Used by the rating engine (Cell 1)", is_input=True)
b.kv("Inside fouling resistance, R_fi", 0.0001, "R_fi", "m²·K/W", "", is_input=True)
b.kv("Outside fouling resistance, R_fo", 0.0002, "R_fo", "m²·K/W", "", is_input=True)
b.kv("Tube roughness", 0.0000015, "tube_roughness", "m", "Smooth drawn tubing", is_input=True)
b.blank()

b.section("GEOMETRY RULES")
b.colheads()
b.kv("Tube pitch ratio, Pt/Do", 1.25, "pitch_ratio", "–", "TEMA triangular pitch", is_input=True)
b.kv("Baffle spacing fraction of shell ID", 0.2, "baffle_frac", "–", "Code names this 'baffle_cut_ratio' but it is really the spacing lB/Ds, not the true TEMA baffle-cut %", is_input=True)
b.kv("Tube wall thickness (search default)", 0.0016, "wall_thickness", "m", "Used while the optimizer searches", is_input=True)
b.blank()

b.section("OPTIMIZER SEARCH BOUNDS  (used by the VBA macro on the 'Optimizer' sheet)")
b.colheads()
b.kv("Tube length, min", 2.0, "L_tube_min", "m", "", is_input=True)
b.kv("Tube length, max", 7.5, "L_tube_max", "m", "", is_input=True)
b.kv("Tube length step", 0.25, "l_tube_step", "m", "", is_input=True)
b.kv("Shell inside diameter, min", 0.200, "D_shell_min", "m", "", is_input=True)
b.kv("Shell inside diameter, max", 2.5, "D_shell_max", "m", "", is_input=True)
b.kv("Shell inside diameter step", 0.05, "d_shell_step", "m", "", is_input=True)
b.kv("Tube outer diameter, min", 0.016, "D_outer_min", "m", "", is_input=True)
b.kv("Tube outer diameter, max", 0.05, "D_outer_max", "m", "", is_input=True)
b.kv("Tube outer diameter step", 0.002, "d_outer_step", "m", "", is_input=True)
b.kv("Max allowable pressure drop (either side)", 70.0, "max_dP_kPa", "kPa", "", is_input=True)
b.kv("Allowed tube-pass counts", "2,4,6,8", "allowed_passes", "–", "Comma list, parsed by the macro", is_input=True)
b.blank()

b.section("KERN RE-RATE — TEMA STANDARD SIZES  (Cell 2 equivalent)")
b.colheads()
b.kv("Standard tube OD", 30.0, "std_od_mm", "mm", "", is_input=True)
b.kv("Standard tube length", 6.1, "std_len_m", "m", "20 ft", is_input=True)
b.kv("Pitch type", "triangular", "pitch_type", "–", "\"triangular\" or \"square\"", is_input=True)
b.kv("Standard wall thickness", 1.6, "std_wall_mm", "mm", "", is_input=True)
b.kv("Hot-side fouling (re-rate)", 0.0001, "fouling_h", "m²·K/W", "", is_input=True)
b.kv("Cold-side fouling (re-rate)", 0.0002, "fouling_c", "m²·K/W", "", is_input=True)
b.kv("Carbon-steel wall conductivity (re-rate)", 50.0, "kern_k_wall", "W/m·K", "Cell 2 hardcodes 50.0, distinct from k_wall=45 above — kept as-is from source", is_input=True)
b.kv("Shell-bundle clearance", 0.038, "clearance_gap", "m", "Fixed 38 mm stand-in for TEMA Fig.12.10 chart (varies 10-90mm by head type/size in the book)", is_input=True)
b.blank()

b.section("NTU / EFFECTIVENESS STAGE  (Cell 3 equivalent)")
b.colheads()
b.kv("NTU-stage cold inlet temp, Tc_in", 25, "ntu_Tc_in", "°C", "NOTE: notebook Cell 3 uses 25 °C here vs 20 °C in Cell 1 — kept as a separate cell, faithful to the source inconsistency", is_input=True)
b.kv("NTU-stage cold outlet temp, Tc_out", 42, "ntu_Tc_out", "°C", "Notebook Cell 3 default", is_input=True)

ws.sheet_view.showGridLines = False

# ======================================================================
# SHEET 2: OPTIMIZER (VBA writes here)
# ======================================================================
ws2 = wb.create_sheet("Optimizer")
b2 = Builder(ws2, "Optimizer")
b2.widths()
b2.title("STAGE 1 — CONSTRAINT-BASED OPTIMIZER  (grid search, run via VBA macro)")
ws2.cell(3, 1, "Click the 'Run Optimizer' button (or Alt+F8 > RunOptimizer) to re-run the exact nested-loop "
                "grid search from optimize_shell_and_tube_cooler() in Python, using the bounds on the Inputs "
                "sheet. It writes the five geometry cells below, which every other sheet in this workbook reads "
                "from live. The values shown now are a pre-loaded example (the notebook's own result) so the "
                "rest of the workbook is usable before you ever run the macro.").font = NOTE
ws2.row_dimensions[3].height = 40
ws2.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
ws2.cell(3,1).alignment = Alignment(wrap_text=True, vertical="top")
b2.row = 6

b2.section("CHOSEN OPTIMAL DESIGN  (macro output — most-compact design meeting all constraints)")
b2.colheads()
b2.kv("Shell inside diameter, D_shell", 0.250, "opt_D_shell", "m", "Example result from the Python optimizer", is_input=True)
b2.kv("Tube outer diameter, D_outer", 0.034, "opt_D_outer", "m", "34 mm", is_input=True)
b2.kv("Tube length, L_tube", 7.5, "opt_L_tube", "m", "", is_input=True)
b2.kv("Number of tube passes, N_passes", 8, "opt_N_passes", "–", "", is_input=True)
b2.kv("Number of tubes, N_t", 32, "opt_N_t", "–", "", is_input=True)
b2.blank()
b2.kv("Optimizer status", "Example (pre-loaded) — click Run Optimizer to recompute", "opt_status", "", "Written by the macro", is_input=True)
b2.kv("Provided area at optimum (m²)", 25.635, "opt_area_hint", "m²", "Informational only, cross-check vs. Rating Engine sheet", is_input=True)
b2.blank(2)
ws2.cell(b2.row, 1, "[Run Optimizer]  <- assign macro RunOptimizer to a Form Control button here (Developer tab > Insert > Button)").font = BOLD
ws2.sheet_view.showGridLines = False


def cell(sym):
    return registry[sym]

# ======================================================================
# SHEET 3: RATING ENGINE (Kern) -- mirrors size_shell_and_tube_cooler()
# ======================================================================
ws3 = wb.create_sheet("Rating Engine (Kern)")
b3 = Builder(ws3, "Rating Engine (Kern)")
b3.widths((34, 18, 10, 60))
b3.title("STAGE 2 — RATING ENGINE  (live formulas, mirrors size_shell_and_tube_cooler())")
ws3.cell(3, 1, "Every cell below is a live formula reading from Inputs and Optimizer. Change any input or the "
               "chosen geometry and this whole sheet recalculates -- this is the Kern's-method check calculation "
               "the optimizer runs internally, made visible.").font = NOTE
ws3.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
ws3.row_dimensions[3].height = 28
ws3.cell(3,1).alignment = Alignment(wrap_text=True, vertical="top")
b3.row = 5

b3.section("GEOMETRY  (from Optimizer sheet)")
b3.colheads()
b3.kv("Shell inside diameter, D_shell_i", f"={cell('opt_D_shell')}", "r_Dshell", "m", "", is_link=True)
b3.kv("Tube outer diameter, D_o", f"={cell('opt_D_outer')}", "r_Do", "m", "", is_link=True)
b3.kv("Tube inner diameter, D_i", f"={cell('opt_D_outer')}-2*{cell('wall_thickness')}", "r_Di", "m",
      "OD - 2 x wall thickness", is_link=True)
b3.kv("Tube length, L_tube", f"={cell('opt_L_tube')}", "r_L", "m", "", is_link=True)
b3.kv("Number of tube passes, N_passes", f"={cell('opt_N_passes')}", "r_Np", "-", "", is_link=True)
b3.kv("Number of tubes, N_t", f"={cell('opt_N_t')}", "r_Nt", "-", "", is_link=True)
b3.blank()

b3.section("ENERGY BALANCE")
b3.colheads()
b3.kv("Heat duty, Q", f"={cell('m_dot_h')}*{cell('cp_h')}*({cell('Th_in')}-{cell('Th_out')})", "r_Q", "W", "m_dot_h . cp_h . (Th_in-Th_out)")
b3.kv("Heat duty, Q", f"={cell('r_Q')}/1000", "r_Q_kW", "kW", "")
b3.kv("Cold utility flow, m_dot_c", f"={cell('r_Q')}/({cell('cp_c')}*({cell('Tc_out')}-{cell('Tc_in')}))", "r_mc", "kg/s", "")
b3.blank()

b3.section("LMTD & MULTI-PASS CORRECTION FACTOR F_T  (Cengel Eq.11-18 / C&R Eq.12.8)")
b3.colheads()
b3.kv("dT1 = Th_in - Tc_out", f"={cell('Th_in')}-{cell('Tc_out')}", "r_dT1", "K", "")
b3.kv("dT2 = Th_out - Tc_in", f"={cell('Th_out')}-{cell('Tc_in')}", "r_dT2", "K", "")
b3.kv("Counter-flow LMTD", f"=IF(ABS({cell('r_dT1')}-{cell('r_dT2')})<0.000001,{cell('r_dT1')},"
      f"IF(OR({cell('r_dT1')}<=0,{cell('r_dT2')}<=0),({cell('r_dT1')}+{cell('r_dT2')})/2,"
      f"({cell('r_dT1')}-{cell('r_dT2')})/LN({cell('r_dT1')}/{cell('r_dT2')})))", "r_lmtd_cf", "K", "Eq.12.4")
b3.kv("delta_Tc = Tc_out - Tc_in", f"={cell('Tc_out')}-{cell('Tc_in')}", "r_dTc", "K", "")
b3.kv("delta_Th = Th_in - Th_out", f"={cell('Th_in')}-{cell('Th_out')}", "r_dTh", "K", "")
b3.kv("Th_in - Tc_in", f"={cell('Th_in')}-{cell('Tc_in')}", "r_dThTcin", "K", "")
b3.kv("R (capacity ratio)", f"=IF(ABS({cell('r_dTc')})>0.000001,{cell('r_dTh')}/{cell('r_dTc')},1000000)", "r_R", "-", "Eq.12.6")
b3.kv("P (=S, thermal efficiency)", f"=IF(ABS({cell('r_dThTcin')})>0.000001,{cell('r_dTc')}/{cell('r_dThTcin')},0)", "r_P", "-", "Eq.12.7")
b3.kv("s = SQRT(R^2+1)", f"=SQRT({cell('r_R')}^2+1)", "r_s", "-", "")
b3.kv("log_arg_num = 1-P", f"=1-{cell('r_P')}", "r_lan", "-", "")
b3.kv("log_arg_den = 1-P.R", f"=1-{cell('r_P')}*{cell('r_R')}", "r_lad", "-", "")
b3.kv("den_term_num = 2-P(R+1-s)", f"=2-{cell('r_P')}*({cell('r_R')}+1-{cell('r_s')})", "r_dtn", "-", "")
b3.kv("den_term_den = 2-P(R+1+s)", f"=2-{cell('r_P')}*({cell('r_R')}+1+{cell('r_s')})", "r_dtd", "-", "")
guard = f"AND({cell('r_lan')}>0,{cell('r_lad')}>0,{cell('r_dtn')}>0,{cell('r_dtd')}>0)"
b3.kv("num_ft = s.LN(log_arg_num/log_arg_den)", f"=IF({guard},{cell('r_s')}*LN({cell('r_lan')}/{cell('r_lad')}),0)", "r_numft", "-", "")
b3.kv("den_ft = (R-1).LN(den_term_num/den_term_den)", f"=IF({guard},({cell('r_R')}-1)*LN({cell('r_dtn')}/{cell('r_dtd')}),1)", "r_denft", "-", "")
ft_formula = (f"=IF(AND({cell('r_P')}>0,{cell('r_P')}<1,{cell('r_R')}>0),"
              f"IF(ABS({cell('r_R')}-1)<0.000001,1,"
              f"IF(AND({guard},ABS({cell('r_denft')})>0.000000001),{cell('r_numft')}/{cell('r_denft')},1)),1)")
b3.kv("F_T (temperature correction factor)", ft_formula, "r_FT", "-", "Guarded exactly as in the notebook; clamped below")
b3.kv("F_T (clamped to [0,1])", f"=IFERROR(IF(OR({cell('r_FT')}<0,{cell('r_FT')}>1),1,{cell('r_FT')}),1)", "r_FT_final", "-", "")
b3.kv("Effective LMTD", f"={cell('r_FT_final')}*{cell('r_lmtd_cf')}", "r_lmtd", "K", "F_T x LMTD_counterflow")
b3.blank()

b3.section("TUBE-SIDE HEAT TRANSFER  (Dittus-Boelter, water)")
b3.colheads()
b3.kv("Single-tube flow area", f"=PI()/4*{cell('r_Di')}^2", "r_atube", "m²", "")
b3.kv("Tubes per pass", f"=MAX(1,{cell('r_Nt')}/{cell('r_Np')})", "r_Ntpp", "-", "")
b3.kv("Tube-side flow area (per pass)", f"={cell('r_Ntpp')}*{cell('r_atube')}", "r_Ac", "m²", "")
b3.kv("Tube-side velocity, u_t", f"={cell('r_mc')}/({cell('rho_c')}*{cell('r_Ac')})", "r_vc", "m/s", "")
b3.kv("Tube-side Reynolds number, Re_c", f"={cell('rho_c')}*{cell('r_vc')}*{cell('r_Di')}/{cell('mu_c')}", "r_Rec", "-", "")
b3.kv("Tube-side Prandtl number, Pr_c", f"={cell('cp_c')}*{cell('mu_c')}/{cell('k_c')}", "r_Prc", "-", "")
b3.kv("Tube-side Nusselt number, Nu_c", f"=IF({cell('r_Rec')}<2300,3.66,0.023*{cell('r_Rec')}^0.8*{cell('r_Prc')}^0.4)", "r_Nuc", "-", "n=0.4 (heating)")
b3.kv("Tube-side coefficient, h_i", f"={cell('r_Nuc')}*{cell('k_c')}/{cell('r_Di')}", "r_hi", "W/m²K", "")
b3.blank()

b3.section("SHELL-SIDE HEAT TRANSFER  (Kern, 30 deg triangular pitch, oil)")
b3.colheads()
b3.kv("Tube pitch, P_t", f"={cell('pitch_ratio')}*{cell('r_Do')}", "r_Pt", "m", "")
b3.kv("Clearance between tubes", f"={cell('r_Pt')}-{cell('r_Do')}", "r_Cclear", "m", "")
b3.kv("Baffle spacing, B", f"={cell('baffle_frac')}*{cell('r_Dshell')}", "r_B", "m", "")
b3.kv("Shell-side cross-flow area", f"={cell('r_Dshell')}*{cell('r_Cclear')}*{cell('r_B')}/{cell('r_Pt')}", "r_Ah", "m²", "")
b3.kv("Shell-side hydraulic diameter, D_e", f"=(3.464*{cell('r_Pt')}^2-PI()*{cell('r_Do')}^2)/(PI()*{cell('r_Do')})", "r_De", "m", "")
b3.kv("Shell-side velocity, u_s", f"={cell('m_dot_h')}/({cell('rho_h')}*{cell('r_Ah')})", "r_vh", "m/s", "")
b3.kv("Shell-side Reynolds number, Re_h", f"={cell('rho_h')}*{cell('r_vh')}*{cell('r_De')}/{cell('mu_h')}", "r_Reh", "-", "")
b3.kv("Shell-side Prandtl number, Pr_h", f"={cell('cp_h')}*{cell('mu_h')}/{cell('k_h')}", "r_Prh", "-", "")
b3.kv("Shell-side Nusselt number, Nu_h", f"=IF({cell('r_Reh')}<2300,3.66,0.023*{cell('r_Reh')}^0.8*{cell('r_Prh')}^0.3)", "r_Nuh", "-", "n=0.3 (cooling)")
b3.kv("Shell-side coefficient, h_o", f"={cell('r_Nuh')}*{cell('k_h')}/{cell('r_De')}", "r_ho", "W/m²K", "")
b3.blank()

b3.section("OVERALL HEAT TRANSFER COEFFICIENT & AREA")
b3.colheads()
b3.kv("Tube-side conv. resistance term", f"={cell('r_Do')}/({cell('r_Di')}*{cell('r_hi')})", "r_tct", "-", "")
b3.kv("Wall resistance term", f"={cell('r_Do')}*LN({cell('r_Do')}/{cell('r_Di')})/(2*{cell('k_wall')})", "r_twall", "-", "")
b3.kv("Shell-side conv. resistance term", f"=1/{cell('r_ho')}", "r_tcs", "-", "")
b3.kv("1/U_o (sum of resistances)", f"={cell('r_tcs')}+{cell('R_fo')}+{cell('r_twall')}+{cell('r_tct')}+{cell('R_fi')}*({cell('r_Do')}/{cell('r_Di')})", "r_invU", "-", "")
b3.kv("Overall coefficient, U_o", f"=1/{cell('r_invU')}", "r_U", "W/m²K", "")
b3.kv("Required surface area", f"=IF({cell('r_lmtd')}>0,{cell('r_Q')}/({cell('r_U')}*{cell('r_lmtd')}),1000000)", "r_Areq", "m²", "")
b3.kv("Provided surface area", f"={cell('r_Nt')}*PI()*{cell('r_Do')}*{cell('r_L')}", "r_Aprov", "m²", "")
b3.kv("Design margin (provided/required)", f"={cell('r_Aprov')}/{cell('r_Areq')}", "r_margin", "-", "Must be >= 1")
b3.blank()

b3.section("PRESSURE DROP")
b3.colheads()
b3.kv("Tube friction factor, f_tube", f"=IF({cell('r_Rec')}<2300,64/{cell('r_Rec')},"
      f"(-1.8*LOG10((({cell('tube_roughness')}/{cell('r_Di')})/3.7)^1.11+6.9/{cell('r_Rec')}))^-2)", "r_ftube", "-", "Laminar / Haaland")
b3.kv("Tube-side dP, friction component", f"={cell('r_Np')}*{cell('r_ftube')}*({cell('r_L')}/{cell('r_Di')})*(0.5*{cell('rho_c')}*{cell('r_vc')}^2)", "r_dPt1", "Pa", "")
b3.kv("Tube-side dP, return-loss component", f"={cell('r_Np')}*2.5*(0.5*{cell('rho_c')}*{cell('r_vc')}^2)", "r_dPt2", "Pa", "2.5 vel. heads/pass, C&R Eq.12.20 (Frank)")
b3.kv("Tube-side pressure drop", f"=({cell('r_dPt1')}+{cell('r_dPt2')})/1000", "r_dPtube", "kPa", "")
b3.kv("Number of baffles", f"=IF({cell('r_L')}>{cell('r_B')},INT({cell('r_L')}/{cell('r_B')})-1,0)", "r_Nbaf", "-", "")
b3.kv("Number of crossings", f"={cell('r_Nbaf')}+1", "r_Ncross", "-", "")
b3.kv("Shell friction factor, f_shell", f"=IF({cell('r_Reh')}<2300,64/{cell('r_Reh')},0.316*{cell('r_Reh')}^-0.25)", "r_fshell", "-", "")
b3.kv("Shell-side pressure drop", f"={cell('r_fshell')}*({cell('r_Dshell')}/{cell('r_De')})*(0.5*{cell('rho_h')}*{cell('r_vh')}^2)*{cell('r_Ncross')}/1000", "r_dPshell", "kPa", "")
b3.blank()
b3.kv("Meets pressure-drop constraint?", f'=IF(AND({cell("r_dPtube")}<={cell("max_dP_kPa")},{cell("r_dPshell")}<={cell("max_dP_kPa")}),"YES","NO")', "r_dp_ok", "-", "")
b3.kv("Meets area constraint?", f'=IF({cell("r_Aprov")}>={cell("r_Areq")},"YES","NO")', "r_area_ok", "-", "")
ws3.sheet_view.showGridLines = False

# ======================================================================
# SHEET 4: KERN RE-RATE (TEMA STANDARD) -- mirrors re_rate_exchanger_kern_method()
# ======================================================================
ws4 = wb.create_sheet("Kern Re-Rate")
b4 = Builder(ws4, "Kern Re-Rate")
b4.widths((34, 18, 10, 62))
b4.title("STAGE 3 — KERN RE-RATE TO TEMA STANDARD SIZES  (live formulas)")
ws4.cell(3, 1, "Re-rates the design onto standard tube OD/length using Kern's shell-side correlation and the "
               "pass-count-aware bundle-diameter constants (Table 12.4). One formula here (tube-side flow area) "
               "was corrected from the source notebook -- see the flagged note below.").font = NOTE
ws4.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
ws4.row_dimensions[3].height = 28
ws4.cell(3,1).alignment = Alignment(wrap_text=True, vertical="top")
b4.row = 5

b4.section("BUNDLE-DIAMETER CONSTANTS  (C&R Table 12.4, by pitch type & pass count)")
ws4.cell(b4.row,1,"Passes").font = BOLD
ws4.cell(b4.row,2,"K1 (tri)").font = BOLD
ws4.cell(b4.row,3,"n1 (tri)").font = BOLD
ws4.cell(b4.row,4,"K1 (sq)").font = BOLD
ws4.cell(b4.row,5,"n1 (sq)").font = BOLD
b4.row += 1
table_start_row = b4.row
table_data = [
    (1, 0.319, 2.142, 0.215, 2.207),
    (2, 0.249, 2.207, 0.156, 2.291),
    (4, 0.175, 2.285, 0.158, 2.263),
    (6, 0.0743, 2.499, 0.0402, 2.617),
    (8, 0.0365, 2.675, 0.0331, 2.643),
]
for passes, k1t, n1t, k1s, n1s in table_data:
    ws4.cell(b4.row, 1, passes)
    ws4.cell(b4.row, 2, k1t)
    ws4.cell(b4.row, 3, n1t)
    ws4.cell(b4.row, 4, k1s)
    ws4.cell(b4.row, 5, n1s)
    for col in range(1, 6):
        ws4.cell(b4.row, col).border = BORDER
        ws4.cell(b4.row, col).font = BLACK
    b4.row += 1
table_end_row = b4.row - 1
pass_rng = f"$A${table_start_row}:$A${table_end_row}"
k1t_rng = f"$B${table_start_row}:$B${table_end_row}"
n1t_rng = f"$C${table_start_row}:$C${table_end_row}"
k1s_rng = f"$D${table_start_row}:$D${table_end_row}"
n1s_rng = f"$E${table_start_row}:$E${table_end_row}"
b4.blank()

b4.section("TARGETS FROM RATING ENGINE / OPTIMIZER")
b4.colheads()
b4.kv("Target area (from Rating Engine)", f"={cell('r_Aprov')}", "k_Atarget", "m²", "Provided Surface Area from Stage 2", is_link=True)
b4.kv("Number of tube passes", f"={cell('opt_N_passes')}", "k_Np", "-", "", is_link=True)
b4.blank()

b4.section("GEOMETRY CONVERSIONS TO SI")
b4.colheads()
b4.kv("Standard tube OD, d_o", f"={cell('std_od_mm')}/1000", "k_do", "m", "")
b4.kv("Standard length, L", f"={cell('std_len_m')}", "k_L", "m", "", is_link=True)
b4.kv("Wall thickness, t_wall", f"={cell('std_wall_mm')}/1000", "k_twall", "m", "")
b4.kv("Standard tube ID, d_i", f"={cell('k_do')}-2*{cell('k_twall')}", "k_di", "m", "")
b4.blank()

b4.section("TUBE COUNT & SURFACE AREA")
b4.colheads()
b4.kv("Area per standard tube", f"=PI()*{cell('k_do')}*{cell('k_L')}", "k_apt", "m²", "")
b4.kv("Number of tubes, n_tubes", f"=CEILING({cell('k_Atarget')}/{cell('k_apt')},1)", "k_nt", "-", "")
b4.kv("Provided surface area", f"={cell('k_nt')}*{cell('k_apt')}", "k_Aprov", "m²", "")
b4.blank()

b4.section("BUNDLE & SHELL DIAMETER  (pass-aware K1, n1 -- FIXED from hardcoded 6-pass values)")
b4.colheads()
k1_formula = (f'=IFERROR(IF(LOWER({cell("pitch_type")})="triangular",'
              f'INDEX({k1t_rng},MATCH({cell("k_Np")},{pass_rng},0)),'
              f'INDEX({k1s_rng},MATCH({cell("k_Np")},{pass_rng},0))),'
              f'IF(LOWER({cell("pitch_type")})="triangular",INDEX({k1t_rng},2),INDEX({k1s_rng},2)))')
n1_formula = (f'=IFERROR(IF(LOWER({cell("pitch_type")})="triangular",'
              f'INDEX({n1t_rng},MATCH({cell("k_Np")},{pass_rng},0)),'
              f'INDEX({n1s_rng},MATCH({cell("k_Np")},{pass_rng},0))),'
              f'IF(LOWER({cell("pitch_type")})="triangular",INDEX({n1t_rng},2),INDEX({n1s_rng},2)))')
b4.kv("K1 (looked up by pass count)", k1_formula, "k_K1", "-", "Falls back to the 2-pass row if N_passes isn't 1/2/4/6/8")
b4.kv("n1 (looked up by pass count)", n1_formula, "k_n1", "-", "")
b4.kv("Bundle diameter, Db", f"={cell('k_do')}*({cell('k_nt')}/{cell('k_K1')})^(1/{cell('k_n1')})", "k_Db", "m",
      "Eq.12.3b. Validated vs C&R Example 12.1 (918 tubes/2 pass -> 0.826 m)")
b4.kv("Shell-bundle clearance", f"={cell('clearance_gap')}", "k_clear", "m", "Fixed stand-in for Fig.12.10 chart", is_link=True)
b4.kv("Raw shell diameter", f"={cell('k_Db')}+{cell('k_clear')}", "k_Draw", "m", "")
b4.kv("Standard shell ID", f"=CEILING({cell('k_Draw')}/0.05,1)*0.05", "k_Dshell", "m", "Rounded up to nearest 50 mm")
b4.blank()

b4.section("TUBE-SIDE  (water)")
b4.colheads()
b4.kv("Tube pitch", f"={cell('pitch_ratio')}*{cell('std_od_mm')}/1000", "k_Pt", "m", "")
b4.kv("Baffle spacing", f"=0.4*{cell('k_Dshell')}", "k_Bspace", "m",
      "Notebook Cell 2 default (0.4) differs from Cell 1's search value (0.2) -- kept faithful to source; edit if you want them to match")
b4.kv("Tubes per pass", f"={cell('k_nt')}/{cell('k_Np')}", "k_ntpp", "-",
      "FIX: notebook Cell 2 used N_passes . tube_area (ignoring n_tubes entirely) for the tube-side flow area -- "
      "a real bug, ~40% off in the validation case. Corrected here to n_tubes/N_passes, matching Cell 1's own logic.")
b4.kv("Tube-side flow area", f"={cell('k_ntpp')}*PI()/4*{cell('k_di')}^2", "k_Atube", "m²", "")
b4.kv("Tube-side velocity, u_t", f"={cell('r_mc')}/({cell('rho_c')}*{cell('k_Atube')})", "k_vtube", "m/s", "")
b4.kv("Tube-side Reynolds number", f"={cell('rho_c')}*{cell('k_vtube')}*{cell('k_di')}/{cell('mu_c')}", "k_Retube", "-", "")
b4.kv("Tube-side Prandtl number", f"={cell('cp_c')}*{cell('mu_c')}/{cell('k_c')}", "k_Prtube", "-", "")
b4.kv("Tube-side Nusselt number", f"=IF({cell('k_Retube')}>2300,0.023*{cell('k_Retube')}^0.8*{cell('k_Prtube')}^0.4,3.66)", "k_Nutube", "-", "")
b4.kv("Recalculated h_i", f"={cell('k_Nutube')}*{cell('k_c')}/{cell('k_di')}", "k_hi", "W/m²K", "")
b4.blank()

b4.section("SHELL-SIDE  (Kern's method, oil -- unchanged, established correlation)")
b4.colheads()
b4.kv("Clearance between tubes", f"={cell('k_Pt')}-{cell('k_do')}", "k_ctube", "m", "")
b4.kv("Shell-side cross-flow area", f"={cell('k_Dshell')}*{cell('k_ctube')}*{cell('k_Bspace')}/{cell('k_Pt')}", "k_Ashell", "m²", "")
b4.kv("Shell-side mass velocity, G_s", f"={cell('m_dot_h')}/{cell('k_Ashell')}", "k_Gshell", "kg/m²s", "")
b4.kv("Kern equivalent diameter, D_e", f"=(4*((SQRT(3)/4)*{cell('k_Pt')}^2-(PI()/8)*{cell('k_do')}^2))/(PI()*{cell('k_do')}/2)", "k_De", "m", "Triangular pitch, Eq.12.23")
b4.kv("Shell-side Reynolds number", f"={cell('k_Gshell')}*{cell('k_De')}/{cell('mu_h')}", "k_Reshell", "-", "")
b4.kv("Shell-side Prandtl number", f"={cell('cp_h')}*{cell('mu_h')}/{cell('k_h')}", "k_Prshell", "-", "")
b4.kv("Kern shell-side h_o", f"=IF({cell('k_Reshell')}>100,0.36*({cell('k_h')}/{cell('k_De')})*{cell('k_Reshell')}^0.55*{cell('k_Prshell')}^(1/3),1.0*({cell('k_h')}/{cell('k_De')}))", "k_ho", "W/m²K",
      "Kern's original closed-form correlation -- kept as-is per your direction. Uses the single canonical "
      "k_h/mu_h from Inputs; the notebook's Cell 2 hardcoded slightly different values (k_h=0.147, mu_h=0.0116 "
      "vs Cell 1's 0.1402/0.0116192) in its execution call -- consolidated here rather than reproduced, "
      "~3% effect on this h_o")
b4.blank()

b4.section("OVERALL U")
b4.colheads()
b4.kv("Wall resistance", f"={cell('k_do')}*LN({cell('k_do')}/{cell('k_di')})/(2*{cell('kern_k_wall')})", "k_twallres", "-", "")
b4.kv("1/U (sum of resistances)", f"=1/{cell('k_ho')}+{cell('fouling_h')}+{cell('k_twallres')}+({cell('k_do')}/{cell('k_di')})*{cell('fouling_c')}+({cell('k_do')}/({cell('k_di')}*{cell('k_hi')}))", "k_invU", "-", "")
b4.kv("Recalculated overall U", f"=1/{cell('k_invU')}", "k_U", "W/m²K", "")
ws4.sheet_view.showGridLines = False

# ======================================================================
# SHEET 5: NTU EFFECTIVENESS -- mirrors evaluate_tema_exchanger_effectiveness()
# ======================================================================
ws5 = wb.create_sheet("NTU Effectiveness")
b5 = Builder(ws5, "NTU Effectiveness")
b5.widths((34, 18, 10, 62))
b5.title("STAGE 4 — NTU / EFFECTIVENESS  (live formulas, 1 shell pass / even tube passes)")
ws5.cell(3, 1, "Not covered by C&R Ch.12 itself (the book defers to Vol.1) -- this is the standard "
               "1-shell-pass effectiveness-NTU relation (Incropera), matching the notebook's Cell 3.").font = NOTE
ws5.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
ws5.row_dimensions[3].height = 26
ws5.cell(3,1).alignment = Alignment(wrap_text=True, vertical="top")
b5.row = 5

b5.section("CAPACITY RATES")
b5.colheads()
b5.kv("Hot fluid capacity rate, C_h", f"={cell('m_dot_h')}*{cell('cp_h')}", "n_Ch", "W/K", "")
b5.kv("Cold fluid capacity rate, C_c", f"={cell('r_mc')}*{cell('cp_c')}", "n_Cc", "W/K", "m_dot_c carried from Stage 2")
b5.kv("Minimum capacity rate, C_min", f"=MIN({cell('n_Ch')},{cell('n_Cc')})", "n_Cmin", "W/K", "")
b5.kv("Maximum capacity rate, C_max", f"=MAX({cell('n_Ch')},{cell('n_Cc')})", "n_Cmax", "W/K", "")
b5.kv("Capacity ratio, C_r", f"={cell('n_Cmin')}/{cell('n_Cmax')}", "n_Cr", "-", "")
b5.blank()

b5.section("NTU & EFFECTIVENESS")
b5.colheads()
b5.kv("Overall U (from Kern re-rate)", f"={cell('k_U')}", "n_U", "W/m²K", "", is_link=True)
b5.kv("Provided area (from Kern re-rate)", f"={cell('k_Aprov')}", "n_A", "m²", "", is_link=True)
b5.kv("Number of transfer units, NTU", f"={cell('n_U')}*{cell('n_A')}/{cell('n_Cmin')}", "n_NTU", "-", "")
b5.kv("gamma = SQRT(1+Cr^2)", f"=SQRT(1+{cell('n_Cr')}^2)", "n_gamma", "-", "")
b5.kv("Exponential term, EXP(-NTU.gamma)", f"=EXP(-{cell('n_NTU')}*{cell('n_gamma')})", "n_exp", "-", "")
eps_formula = (f"=IF(ABS(1-{cell('n_exp')})<0.000000001,1/(1+{cell('n_Cr')}),"
               f"2/(1+{cell('n_Cr')}+{cell('n_gamma')}*(1+{cell('n_exp')})/(1-{cell('n_exp')})))")
b5.kv("Predicted effectiveness, epsilon", eps_formula, "n_eps", "-", "1-shell-pass / even tube-pass relation")
b5.blank()

b5.section("ENERGY CHECK  (NOTE: uses the separate NTU-stage Tc_in, per source notebook Cell 3)")
b5.colheads()
b5.kv("Actual heat duty, Q_actual", f"={cell('r_Q_kW')}", "n_Qact_kW", "kW", "Carried from Stage 2", is_link=True)
b5.kv("Actual heat duty, Q_actual", f"={cell('n_Qact_kW')}*1000", "n_Qact_W", "W", "")
b5.kv("Max possible heat transfer, Q_max", f"={cell('n_Cmin')}*({cell('Th_in')}-{cell('ntu_Tc_in')})", "n_Qmax", "W", "Uses NTU-stage Tc_in, not the Stage-2 Tc_in -- faithful to the notebook's own inconsistency")
b5.kv("Actual operational effectiveness ratio", f"=IF({cell('n_Qmax')}>0,{cell('n_Qact_W')}/{cell('n_Qmax')},0)", "n_eps_actual", "-", "Q_actual / Q_max")
ws5.sheet_view.showGridLines = False

# ======================================================================
# SHEET 0: READ ME (inserted first)
# ======================================================================
ws0 = wb.create_sheet("Read Me", 0)
b0 = Builder(ws0, "Read Me")
b0.widths((100, 1, 1, 1))
b0.title("SHELL & TUBE HX DESIGN — EXCEL PORT OF CIO_Part_1.ipynb")
b0.row = 3

lines = [
    ("How this workbook is organised", True),
    ("1. Inputs — every editable number (blue/yellow cells). Nothing else on this sheet is a formula.", False),
    ("2. Optimizer — the 5 geometry values chosen by the grid-search optimizer (D_shell, D_outer, L_tube, "
     "N_passes, N_t). Pre-loaded with the notebook's own result; click Run Optimizer (VBA) to recompute for "
     "whatever is on the Inputs sheet.", False),
    ("3. Rating Engine (Kern) — live formulas, mirrors size_shell_and_tube_cooler(). This is the full Kern "
     "check calculation: LMTD/F_T, tube-side Dittus-Boelter, shell-side Kern correlation, U, area, both "
     "pressure drops. Recalculates the instant you change an input or the chosen geometry.", False),
    ("4. Kern Re-Rate — live formulas, mirrors re_rate_exchanger_kern_method(). Standardises onto real tube "
     "OD/length and TEMA shell sizes.", False),
    ("5. NTU Effectiveness — live formulas, mirrors evaluate_tema_exchanger_effectiveness().", False),
    ("6. VBA Module — read-only listing of the macro source, for reference/audit.", False),
    ("", False),
    ("Why VBA only runs the optimizer", True),
    ("Excel has no native nested-loop grid search. Only Stage 1 (the four-level loop over shell diameter, "
     "tube OD, pass count and tube length) is implemented in VBA -- a line-for-line translation of "
     "optimize_shell_and_tube_cooler(), same bounds, same step sizes. Everything downstream of that (Stages "
     "2-4) is native Excel formulas, not macro code, so the workbook is fully auditable and works even with "
     "macros disabled (using the pre-loaded example geometry).", False),
    ("", False),
    ("To enable the macro", True),
    ("This ships as a plain .xlsx -- openpyxl (which built this file) cannot author a real embedded VBA "
     "project, so a fabricated .xlsm would have a mismatched, Excel-rejected file format. To add the macro: "
     "open this .xlsx in Excel, press Alt+F11, right-click VBAProject > Import File, and import "
     "ShellTubeOptimizer.bas (included alongside this workbook). Then File > Save As > pick 'Excel Macro-Enabled "
     "Workbook (.xlsm)' -- Excel itself now writes a real vbaProject.bin. Then Developer tab > Insert > Button, "
     "draw it on the Optimizer sheet, and assign macro RunOptimizer. Or just press Alt+F8 and run RunOptimizer "
     "directly, no button needed.", False),
    ("", False),
    ("Three corrections made vs. the source notebook (all flagged inline where they occur)", True),
    ("1. Kern Re-Rate: bundle-diameter constants K1/n1 are now looked up by pass count (Table 12.4) instead "
     "of being hardcoded to the 6-pass row -- validated against C&R Example 12.1 (826 mm).", False),
    ("2. Rating Engine: tube-side return-loss allowance uses 2.5 velocity heads/pass (C&R Eq.12.20, Frank's "
     "correction) instead of Kern's original 4.0.", False),
    ("3. Kern Re-Rate: tube-side flow area now uses tubes-per-pass = n_tubes/N_passes (matching the Rating "
     "Engine's own logic), instead of the notebook's n_passes . single-tube-area, which silently ignored "
     "n_tubes altogether -- a newly found bug, ~40% off in the validation case. Not yet applied back to the "
     "Python notebook; flag if you want that patched too.", False),
    ("", False),
    ("Left unchanged, per your direction", True),
    ("Dittus-Boelter for tube-side h_i, and Kern's original closed-form shell-side correlation "
     "(h_o = 0.36 . Re^0.55 . Pr^(1/3)) for shell-side h_o. Both are established methods.", False),
]
for text, is_head in lines:
    c = ws0.cell(b0.row, 1, text)
    c.font = SECTION if is_head else Font(name=FONT_NAME, size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.row_dimensions[b0.row].height = 15 if not text else (18 if is_head else max(15, 15*(len(text)//110+1)))
    b0.row += 1
ws0.sheet_view.showGridLines = False

# ======================================================================
# SHEET: VBA MODULE (reference listing)
# ======================================================================
ws6 = wb.create_sheet("VBA Module")
ws6.column_dimensions["A"].width = 130
ws6.cell(1,1,"REFERENCE LISTING — import ShellTubeOptimizer.bas into the VBA editor rather than retyping this.").font = BOLD
ws6.sheet_view.showGridLines = False
mono = Font(name="Consolas", size=9)

VBA_SOURCE = r'''Option Explicit

' ============================================================
' ShellTubeOptimizer.bas
' Line-for-line translation of optimize_shell_and_tube_cooler()
' from CIO_Part_1.ipynb. Reads bounds from the Inputs sheet,
' writes the winning design to the Optimizer sheet.
' ============================================================

Private Function MilOilCp() As Double
    MilOilCp = 1903.1
End Function

Function LMTD_CF(Th_in As Double, Th_out As Double, Tc_in As Double, Tc_out As Double) As Double
    Dim dT1 As Double, dT2 As Double
    dT1 = Th_in - Tc_out
    dT2 = Th_out - Tc_in
    If Abs(dT1 - dT2) < 0.000001 Then
        LMTD_CF = dT1
    ElseIf dT1 <= 0 Or dT2 <= 0 Then
        LMTD_CF = (dT1 + dT2) / 2
    Else
        LMTD_CF = (dT1 - dT2) / Log(dT1 / dT2)
    End If
End Function

Function CorrectionFactorFT(Th_in As Double, Th_out As Double, Tc_in As Double, Tc_out As Double) As Double
    Dim dTc As Double, dTh As Double, dThTcin As Double
    Dim R As Double, P As Double, s As Double
    Dim lan As Double, lad As Double, dtn As Double, dtd As Double
    Dim numFt As Double, denFt As Double, ft As Double

    dTc = Tc_out - Tc_in
    dTh = Th_in - Th_out
    dThTcin = Th_in - Tc_in
    R = IIf(Abs(dTc) > 0.000001, dTh / dTc, 1000000#)
    P = IIf(Abs(dThTcin) > 0.000001, dTc / dThTcin, 0#)

    ft = 1#
    If P > 0 And P < 1 And R > 0 Then
        If Abs(R - 1) < 0.000001 Then
            ft = 1#
        Else
            s = Sqr(R ^ 2 + 1)
            lan = 1 - P
            lad = 1 - P * R
            dtn = 2 - P * (R + 1 - s)
            dtd = 2 - P * (R + 1 + s)
            If lan > 0 And lad > 0 And dtn > 0 And dtd > 0 Then
                numFt = s * Log(lan / lad)
                denFt = (R - 1) * Log(dtn / dtd)
                If Abs(denFt) > 0.000000001 Then ft = numFt / denFt
            End If
        End If
    End If
    If ft < 0 Or ft > 1 Then ft = 1#
    CorrectionFactorFT = ft
End Function

' Mirrors size_shell_and_tube_cooler(): returns Area_required, Area_provided,
' dP_tube (kPa), dP_shell (kPa) for one candidate geometry.
Sub RateDesign(Th_in As Double, Th_out As Double, mDotH As Double, _
                Tc_in As Double, Tc_out As Double, cpC As Double, rhoC As Double, muC As Double, kC As Double, _
                rhoH As Double, cpH As Double, kH As Double, muH As Double, _
                Di As Double, Do_ As Double, Dshell As Double, Ltube As Double, kWall As Double, _
                Rfi As Double, Rfo As Double, pitchRatio As Double, baffleFrac As Double, roughness As Double, _
                Nt As Long, Np As Long, _
                ByRef areaReq As Double, ByRef areaProv As Double, ByRef dPtube As Double, ByRef dPshell As Double)

    Dim Q As Double, mDotC As Double, lmtdCf As Double, ft As Double, lmtd As Double
    Dim aTube As Double, NtPerPass As Double, Ac As Double, vC As Double, ReC As Double, PrC As Double, NuC As Double, hi As Double
    Dim Pt As Double, Cclear As Double, Bbaf As Double, Ah As Double, De As Double
    Dim vH As Double, ReH As Double, PrH As Double, NuH As Double, ho As Double
    Dim invU As Double, U As Double
    Dim fTube As Double, relRough As Double, Nbaf As Long, Ncross As Long, fShell As Double

    Q = mDotH * cpH * (Th_in - Th_out)
    mDotC = Q / (cpC * (Tc_out - Tc_in))

    lmtdCf = LMTD_CF(Th_in, Th_out, Tc_in, Tc_out)
    ft = CorrectionFactorFT(Th_in, Th_out, Tc_in, Tc_out)
    lmtd = ft * lmtdCf

    aTube = Application.WorksheetFunction.Pi() / 4 * Di ^ 2
    NtPerPass = Application.WorksheetFunction.Max(1, Nt / Np)
    Ac = NtPerPass * aTube
    vC = mDotC / (rhoC * Ac)
    ReC = rhoC * vC * Di / muC
    PrC = cpC * muC / kC
    If ReC < 2300 Then NuC = 3.66 Else NuC = 0.023 * ReC ^ 0.8 * PrC ^ 0.4
    hi = NuC * kC / Di

    Pt = pitchRatio * Do_
    Cclear = Pt - Do_
    Bbaf = baffleFrac * Dshell
    Ah = Dshell * Cclear * Bbaf / Pt
    De = (3.464 * Pt ^ 2 - Application.WorksheetFunction.Pi() * Do_ ^ 2) / (Application.WorksheetFunction.Pi() * Do_)
    vH = mDotH / (rhoH * Ah)
    ReH = rhoH * vH * De / muH
    PrH = cpH * muH / kH
    If ReH < 2300 Then NuH = 3.66 Else NuH = 0.023 * ReH ^ 0.8 * PrH ^ 0.3
    ho = NuH * kH / De

    invU = (1 / ho) + Rfo + (Do_ * Log(Do_ / Di) / (2 * kWall)) + (Do_ / (Di * hi)) + Rfi * (Do_ / Di)
    U = 1 / invU

    If lmtd > 0 Then areaReq = Q / (U * lmtd) Else areaReq = 1E+30
    areaProv = Nt * Application.WorksheetFunction.Pi() * Do_ * Ltube

    If ReC < 2300 Then
        fTube = 64 / ReC
    Else
        relRough = roughness / Di
        fTube = (-1.8 * Application.WorksheetFunction.Log10(((relRough / 3.7) ^ 1.11) + (6.9 / ReC))) ^ -2
    End If
    dPtube = (Np * fTube * (Ltube / Di) * (0.5 * rhoC * vC ^ 2) + Np * 2.5 * (0.5 * rhoC * vC ^ 2)) / 1000

    If Ltube > Bbaf Then Nbaf = Int(Ltube / Bbaf) - 1 Else Nbaf = 0
    Ncross = Nbaf + 1
    If ReH < 2300 Then fShell = 64 / ReH Else fShell = 0.316 * ReH ^ -0.25
    dPshell = fShell * (Dshell / De) * (0.5 * rhoH * vH ^ 2) * Ncross / 1000
End Sub

Sub RunOptimizer()
    Dim wsIn As Worksheet, wsOpt As Worksheet
    Set wsIn = ThisWorkbook.Sheets("Inputs")
    Set wsOpt = ThisWorkbook.Sheets("Optimizer")

    Dim ThIn As Double, ThOut As Double, mDotH As Double
    Dim TcIn As Double, TcOut As Double, cpC As Double, rhoC As Double, muC As Double, kC As Double
    Dim rhoH As Double, cpH As Double, kH As Double, muH As Double
    Dim kWall As Double, Rfi As Double, Rfo As Double, pitchRatio As Double, baffleFrac As Double
    Dim wallThk As Double, roughness As Double
    Dim LtMin As Double, LtMax As Double, LtStep As Double
    Dim DsMin As Double, DsMax As Double, DsStep As Double
    Dim DoMin As Double, DoMax As Double, DoStep As Double
    Dim maxDP As Double
    Dim passesRaw As String, passesArr() As String
    Dim i As Integer

    __INPUT_READS__

    Dim bestAreaProv As Double, found As Boolean
    Dim bestDs As Double, bestDo As Double, bestL As Double, bestNp As Long, bestNt As Long
    bestAreaProv = 1E+30
    found = False

    Dim curDs As Double, curDo As Double, curDi As Double, curL As Double
    Dim pt As Double, approxMaxNt As Long
    Dim pIdx As Integer, np As Long, nt As Long
    Dim areaReq As Double, areaProv As Double, dPtube As Double, dPshell As Double

    curDs = DsMin
    Do While curDs <= DsMax + 0.0000001
        curDo = DoMin
        Do While curDo <= DoMax + 0.0000001
            curDi = curDo - 2 * wallThk
            If curDi > 0 Then
                pt = pitchRatio * curDo
                approxMaxNt = Int(0.75 * ((curDs / pt) ^ 2) * (Application.WorksheetFunction.Pi() / 2))
                If approxMaxNt < 12 Then approxMaxNt = 12

                For pIdx = LBound(passesArr) To UBound(passesArr)
                    np = CLng(Trim(passesArr(pIdx)))
                    nt = np * 2
                    Do While nt <= approxMaxNt
                        curL = LtMin
                        Do While curL <= LtMax + 0.0000001
                            RateDesign ThIn, ThOut, mDotH, TcIn, TcOut, cpC, rhoC, muC, kC, _
                                       rhoH, cpH, kH, muH, curDi, curDo, curDs, curL, kWall, _
                                       Rfi, Rfo, pitchRatio, baffleFrac, roughness, nt, np, _
                                       areaReq, areaProv, dPtube, dPshell

                            If areaProv >= areaReq And dPtube <= maxDP And dPshell <= maxDP Then
                                If areaProv < bestAreaProv Then
                                    bestAreaProv = areaProv
                                    bestDs = curDs: bestDo = curDo: bestL = curL: bestNp = np: bestNt = nt
                                    found = True
                                End If
                            End If
                            curL = curL + LtStep
                        Loop
                        nt = nt + np * 2
                    Loop
                Next pIdx
            End If
            curDo = curDo + DoStep
        Loop
        curDs = curDs + DsStep
    Loop

    If found Then
        __OPT_WRITES_FOUND__
    Else
        __OPT_WRITES_NOTFOUND__
    End If

    MsgBox "Optimizer finished. " & IIf(found, "Best design written to the Optimizer sheet.", "No feasible design found."), vbInformation
End Sub
'''

def rng(key):
    """Registry address ('Inputs!$B$7') -> bare Range() string ('Inputs!B7')."""
    return registry[key].replace("$", "")

input_read_map = [
    ("ThIn", "Th_in"), ("ThOut", "Th_out"), ("mDotH", "m_dot_h"),
    ("rhoH", "rho_h"), ("cpH", "cp_h"), ("kH", "k_h"), ("muH", "mu_h"),
    ("TcIn", "Tc_in"), ("TcOut", "Tc_out"), ("cpC", "cp_c"), ("rhoC", "rho_c"),
    ("muC", "mu_c"), ("kC", "k_c"),
    ("kWall", "k_wall"), ("Rfi", "R_fi"), ("Rfo", "R_fo"), ("roughness", "tube_roughness"),
    ("pitchRatio", "pitch_ratio"), ("baffleFrac", "baffle_frac"), ("wallThk", "wall_thickness"),
    ("LtMin", "L_tube_min"), ("LtMax", "L_tube_max"), ("LtStep", "l_tube_step"),
    ("DsMin", "D_shell_min"), ("DsMax", "D_shell_max"), ("DsStep", "d_shell_step"),
    ("DoMin", "D_outer_min"), ("DoMax", "D_outer_max"), ("DoStep", "d_outer_step"),
    ("maxDP", "max_dP_kPa"), ("passesRaw", "allowed_passes"),
]
input_reads_lines = [f'    {vba_var} = Range("{rng(key)}").Value' for vba_var, key in input_read_map[:-1]]
input_reads_lines.append(f'    passesRaw = Range("{rng("allowed_passes")}").Value')
input_reads_lines.append("    passesArr = Split(passesRaw, \",\")")
input_reads_block = "\n".join(input_reads_lines)

found_writes = "\n        ".join([
    f'wsOpt.Range("{rng("opt_D_shell").split("!")[1]}").Value = bestDs   \' opt_D_shell',
    f'wsOpt.Range("{rng("opt_D_outer").split("!")[1]}").Value = bestDo   \' opt_D_outer',
    f'wsOpt.Range("{rng("opt_L_tube").split("!")[1]}").Value = bestL    \' opt_L_tube',
    f'wsOpt.Range("{rng("opt_N_passes").split("!")[1]}").Value = bestNp   \' opt_N_passes',
    f'wsOpt.Range("{rng("opt_N_t").split("!")[1]}").Value = bestNt   \' opt_N_t',
    f'wsOpt.Range("{rng("opt_status").split("!")[1]}").Value = "Optimal design found " & Format(Now, "yyyy-mm-dd hh:mm")',
    f'wsOpt.Range("{rng("opt_area_hint").split("!")[1]}").Value = bestAreaProv',
])
notfound_writes = f'wsOpt.Range("{rng("opt_status").split("!")[1]}").Value = "No configuration met all constraints -- widen bounds or raise max dP"'

VBA_SOURCE = VBA_SOURCE.replace("    __INPUT_READS__", input_reads_block)
VBA_SOURCE = VBA_SOURCE.replace("        __OPT_WRITES_FOUND__", "        " + found_writes)
VBA_SOURCE = VBA_SOURCE.replace("        __OPT_WRITES_NOTFOUND__", "        " + notfound_writes)

for i, line in enumerate(VBA_SOURCE.strip("\n").split("\n"), start=3):
    c = ws6.cell(i, 1, line)
    c.font = mono

# ======================================================================
# Sheet order + freeze panes + save
# ======================================================================
for s in [ws0, ws, ws2, ws3, ws4, ws5, ws6]:
    s.freeze_panes = "A6"
ws.freeze_panes = "A6"

out_path = "/tmp/claude-0/-home-user-Claude/485fa6e6-dda5-5a6f-befa-fd08ce7dd6bb/scratchpad/xlsx_build/Shell_and_Tube_HX_Design.xlsx"
wb.save(out_path)

bas_path = "/tmp/claude-0/-home-user-Claude/485fa6e6-dda5-5a6f-befa-fd08ce7dd6bb/scratchpad/xlsx_build/ShellTubeOptimizer.bas"
with open(bas_path, "w") as f:
    f.write(VBA_SOURCE.strip("\n") + "\n")

print("Saved:", out_path)
print("Saved:", bas_path)
print("Registry keys:", len(registry))


